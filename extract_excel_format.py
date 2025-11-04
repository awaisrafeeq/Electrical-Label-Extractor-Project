#!/usr/bin/env python3
"""
Extract formatting rules from Label Project.xlsx
"""
import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('Label Project.xlsx')

print("="*80)
print("EXCEL FORMATTING ANALYSIS - Label Project.xlsx")
print("="*80)

# Analyze all sheets
for sheet_name in wb.sheetnames[:15]:  # First 15 sheets
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print("="*80)

    colors_found = defaultdict(list)

    # Check first 50 rows
    for row in range(1, min(50, sheet.max_row + 1)):
        for col in range(1, min(10, sheet.max_column + 1)):
            cell = sheet.cell(row, col)

            if cell.value:
                fill_color = None
                if cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                    fill_color = cell.fill.start_color.rgb

                if fill_color and fill_color != '00000000' and fill_color != 'FFFFFFFF':
                    colors_found[fill_color].append({
                        'row': row,
                        'col': col,
                        'value': str(cell.value)[:50],
                        'font_bold': cell.font.bold,
                        'font_size': cell.font.size
                    })

    # Print color summary
    if colors_found:
        print(f"\nColors found in {sheet_name}:")
        for color, cells in colors_found.items():
            print(f"\n  Color: {color}")
            print(f"  Used in {len(cells)} cells")
            print(f"  Example values: {', '.join([c['value'] for c in cells[:3]])}")
    else:
        print(f"  No special colors found (using default)")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)
