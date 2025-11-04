"""
Excel Exporter
Generates formatted Excel files from extracted label data with color coding
"""
from pathlib import Path
from typing import List, Optional
from datetime import datetime
import pandas as pd
import logging
import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from ..vision_ai.analyzer import LabelData

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export label data to Excel format with color coding"""

    # System color mapping (from LABEL_STYLING_RULES.md)
    SYSTEM_COLORS = {
        'A': 'C4261D',  # Red
        'B': '005197',  # Blue
        'C': 'E15616',  # Orange
        'D': '444785',  # Purple
        'E': 'FEE900',  # Yellow
        'F': 'DEDED8',  # Gray
        'H': '523D2A',  # Brown (House/Emergency)
        'Z': '00B050',  # Green (Spare)
    }

    SPARE_COLOR = '00B050'  # Green
    WARNING_COLOR = 'FEE900'  # Bright Yellow
    DEFAULT_COLOR = 'FFFFFF'  # White

    def __init__(self):
        """Initialize Excel exporter"""
        pass

    def _extract_system(self, device_tag: str) -> str:
        """
        Extract system designation from device tag

        Examples:
            "EDC ATL11 MSBAA110" -> "A"
            "EDC ATL11 MSBAB110" -> "B"
            "EDC ATL11 GENAH100" -> "H"
            "SPARE" -> "Z"
        """
        if not device_tag:
            return None

        # Check if SPARE
        if "SPARE" in device_tag.upper():
            return 'Z'

        # Pattern: Look for equipment code followed by system letters
        # e.g., MSBAA110 -> AA -> A
        # e.g., GENAH100 -> AH -> H
        match = re.search(r'([A-Z]{3,}[A-Z]{1,2})(\d{3,})', device_tag)
        if match:
            code_with_system = match.group(1)
            # Get last two characters before numbers
            if len(code_with_system) >= 4:
                system_chars = code_with_system[-2:]

                # Special handling for H (House systems)
                if 'H' in system_chars:
                    return 'H'

                # Get first character of system designation (AA -> A, AB -> B, etc.)
                return system_chars[0]

        return None

    def _get_cell_color(self, label: LabelData) -> str:
        """
        Get background color for label based on system or special status

        Returns hex color code (without # prefix)
        """
        # Check if SPARE
        if label.is_spare or (label.device_tag and "SPARE" in label.device_tag.upper()):
            return self.SPARE_COLOR

        # Extract system from device tag
        system = self._extract_system(label.device_tag)

        if system and system in self.SYSTEM_COLORS:
            return self.SYSTEM_COLORS[system]

        # Default to white
        return self.DEFAULT_COLOR

    def _apply_color_formatting(self, worksheet, labels: List[LabelData]):
        """
        Apply color coding to Excel cells based on system designation

        Colors the "Full Label", "Line 1", "Line 2", "Line 3", "Line 4" columns
        """
        # Define cell styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )

        # Apply formatting to each row (starting from row 2, after header)
        for idx, label in enumerate(labels, start=2):
            # Get color for this label
            color_hex = self._get_cell_color(label)

            # Create fill pattern
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')

            # Determine text color based on background
            # Use white text for dark backgrounds
            text_color = '000000'  # Black (default)
            if color_hex in ['C4261D', '005197', '444785', '523D2A']:  # Dark colors
                text_color = 'FFFFFF'  # White

            font = Font(name='Arial', size=10, color=text_color)

            # Find column indices for label columns
            # Columns to color: "Full Label", "Line 1", "Line 2", "Line 3", "Line 4"
            header_row = worksheet[1]
            col_indices = {}

            for cell in header_row:
                if cell.value in ["Full Label", "Line 1", "Line 2", "Line 3", "Line 4"]:
                    col_indices[cell.value] = cell.column

            # Apply formatting to label columns
            for col_name, col_idx in col_indices.items():
                cell = worksheet.cell(row=idx, column=col_idx)
                cell.fill = fill
                cell.font = font
                cell.alignment = alignment
                cell.border = thin_border

        # Format header row
        for cell in worksheet[1]:
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        logger.info(f"Applied color formatting to {len(labels)} labels")

    def export_labels(
        self,
        labels: List[LabelData],
        output_path: Path,
        include_metadata: bool = True,
        validation_errors: Optional[List] = None,
        statistics: Optional[any] = None
    ) -> Path:
        """
        Export labels to Excel file

        Args:
            labels: List of LabelData objects
            output_path: Path for output Excel file
            include_metadata: Whether to include metadata sheet

        Returns:
            Path to created Excel file
        """
        logger.info(f"Exporting {len(labels)} labels to Excel...")

        # Prepare data for DataFrame
        rows = []
        for i, label in enumerate(labels, 1):
            # Build label text (2-4 lines)
            label_lines = [label.device_tag]

            # Add feed lines
            if label.primary_from:
                label_lines.append(f"PRIMARY FROM {label.primary_from}")
            if label.alternate_from:
                label_lines.append(f"ALTERNATE FROM {label.alternate_from}")
            elif label.fed_from:
                label_lines.append(f"FED FROM {label.fed_from}")

            # Add specs
            if label.specs:
                label_lines.append(label.specs)

            row = {
                "Index": i,
                "Equipment Type": label.equipment_type,
                "Device Tag": label.device_tag,
                "Line 1": label_lines[0] if len(label_lines) > 0 else "",
                "Line 2": label_lines[1] if len(label_lines) > 1 else "",
                "Line 3": label_lines[2] if len(label_lines) > 2 else "",
                "Line 4": label_lines[3] if len(label_lines) > 3 else "",
                "Full Label": "\n".join(label_lines),
                "Fed From": label.fed_from or label.primary_from or "",
                "Alternate From": label.alternate_from or "",
                "Specs": label.specs or "",
                "Is Spare": "YES" if label.is_spare else "NO",
                "Needs Breaker": "YES" if label.needs_breaker else "NO",
            }
            rows.append(row)

        # Create DataFrame
        df = pd.DataFrame(rows)

        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main labels sheet
            df.to_excel(writer, sheet_name='Labels', index=False)

            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Labels']

            # Apply color coding to labels
            self._apply_color_formatting(worksheet, labels)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet
            if include_metadata:
                self._add_summary_sheet(writer, labels)

            # Add validation errors sheet
            if validation_errors:
                self._add_validation_sheet(writer, validation_errors)

            # Add statistics sheet
            if statistics:
                self._add_statistics_sheet(writer, statistics)

        logger.info(f"Excel file created: {output_path}")
        return output_path

    def _add_summary_sheet(self, writer, labels: List[LabelData]):
        """Add summary/metadata sheet"""
        # Count by equipment type
        equipment_counts = {}
        spare_count = 0

        for label in labels:
            eq_type = label.equipment_type
            equipment_counts[eq_type] = equipment_counts.get(eq_type, 0) + 1
            if label.is_spare:
                spare_count += 1

        # Create summary data
        summary_data = {
            "Metric": [
                "Total Labels",
                "Spare Labels",
                "Equipment Types",
                "Generated Date",
                "Generated Time",
            ],
            "Value": [
                len(labels),
                spare_count,
                len(equipment_counts),
                datetime.now().strftime("%Y-%m-%d"),
                datetime.now().strftime("%H:%M:%S"),
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Equipment type breakdown
        eq_type_data = {
            "Equipment Type": list(equipment_counts.keys()),
            "Count": list(equipment_counts.values())
        }
        eq_df = pd.DataFrame(eq_type_data)
        eq_df = eq_df.sort_values("Count", ascending=False)

        # Write to same sheet, below summary
        eq_df.to_excel(
            writer,
            sheet_name='Summary',
            index=False,
            startrow=len(summary_df) + 3
        )

        # Format summary sheet
        worksheet = writer.sheets['Summary']
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20

    def _add_validation_sheet(self, writer, validation_errors):
        """Add validation errors sheet"""
        if not validation_errors:
            return

        error_data = []
        for error in validation_errors:
            error_data.append({
                "Index": error.label_index + 1,
                "Device Tag": error.device_tag,
                "Error Type": error.error_type,
                "Message": error.message,
                "Severity": error.severity
            })

        df = pd.DataFrame(error_data)
        df.to_excel(writer, sheet_name='Validation Errors', index=False)

        # Format
        worksheet = writer.sheets['Validation Errors']
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 40
        worksheet.column_dimensions['E'].width = 12

    def _add_statistics_sheet(self, writer, statistics):
        """Add statistics sheet"""
        stats_data = []

        # Equipment type counts
        stats_data.append({"Category": "EQUIPMENT TYPE COUNTS", "Value": ""})
        for eq_type, count in statistics.count_by_equipment_type().items():
            stats_data.append({"Category": f"  {eq_type}", "Value": count})

        stats_data.append({"Category": "", "Value": ""})

        # Voltage classes
        stats_data.append({"Category": "VOLTAGE CLASSIFICATION", "Value": ""})
        for voltage_class, count in statistics.count_by_voltage_class().items():
            if count > 0:
                stats_data.append({"Category": f"  {voltage_class}", "Value": count})

        stats_data.append({"Category": "", "Value": ""})

        # Amperage ranges
        stats_data.append({"Category": "AMPERAGE RANGES", "Value": ""})
        for amp_range, count in statistics.count_by_amperage_range().items():
            if count > 0:
                stats_data.append({"Category": f"  {amp_range}", "Value": count})

        stats_data.append({"Category": "", "Value": ""})

        # Spare count
        spare_count = statistics.count_spare_labels()
        stats_data.append({"Category": "SPARE LABELS", "Value": spare_count})

        df = pd.DataFrame(stats_data)
        df.to_excel(writer, sheet_name='Statistics', index=False)

        # Format
        worksheet = writer.sheets['Statistics']
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 15

    def export_by_equipment_type(
        self,
        labels: List[LabelData],
        output_dir: Path
    ) -> List[Path]:
        """
        Export labels grouped by equipment type into separate Excel files

        Args:
            labels: List of LabelData objects
            output_dir: Directory for output files

        Returns:
            List of created file paths
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True, parents=True)

        # Group by equipment type
        by_type = {}
        for label in labels:
            eq_type = label.equipment_type
            if eq_type not in by_type:
                by_type[eq_type] = []
            by_type[eq_type].append(label)

        # Export each type
        created_files = []
        for eq_type, type_labels in by_type.items():
            filename = f"labels_{eq_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = output_dir / filename

            self.export_labels(type_labels, file_path, include_metadata=True)
            created_files.append(file_path)

        logger.info(f"Created {len(created_files)} Excel files by equipment type")
        return created_files
